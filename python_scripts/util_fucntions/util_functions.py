import itertools
import math, re, json, csv, os, openpyxl
import pandas as pd
from operator import itemgetter
from itertools import groupby, filterfalse
from pandas.io.json import json_normalize
from datetime import datetime
from tqdm import tqdm
import random
from openpyxl import load_workbook
import torch
import numpy as np


def subtract_lists(large_list: list, small_list: list) -> list:
    return [item for item in large_list if item not in small_list]


def flatten_list(multi_dim_list: list) -> list:
    return list(itertools.chain.from_iterable(multi_dim_list))


def get_unique_list(old_list: list, sort_code=None) -> list:
    # sort_code:
    # 0: no sort
    # 1: sort ascending
    # 2: sort descending
    unique_list = list(set(old_list))
    if sort_code is None or sort_code <= 0:
        return unique_list
    return sorted(unique_list, reverse=sort_code != 1)


def reformat_key(old_key: str) -> str:
    return old_key.lower().replace(' ', '_')


def clean_white_space(old_str: str) -> str:
    return re.sub(r'\s+', ' ', old_str)


def remove_brackets(old_str: str):
    return re.sub(r'\([^()]*\)', '', old_str)


def lower_case_and_clear_white_space(og_string: str, to_regex=False) -> str:
    new_string = og_string.lower().replace(" ", '')
    return re.sub(r'\W+', '', new_string) if to_regex else new_string


def are_strings_the_same(str1: str, str2: str, to_regex=False) -> bool:
    return lower_case_and_clear_white_space(str2, to_regex) == lower_case_and_clear_white_space(str1, to_regex)


def are_strings_similar(str1: str, str2: str, to_regex=False) -> bool:
    str1 = lower_case_and_clear_white_space(str1.replace('KG', '').replace(' kg', '').replace(' kg ', ''), to_regex).replace('imported', '')
    str2 = lower_case_and_clear_white_space(str2.replace('KG', '').replace(' kg', '').replace(' kg ', ''), to_regex).replace('imported', '')
    return str2 in str1


def format_dictionaries(dict_list: list) -> list:
    for data in dict_list:
        modified_keys = {}
        for key, value in data.items():
            modified_keys[reformat_key(key)] = value
        data.clear()
        data.update(modified_keys)
    return dict_list


def dictionary_has_nan(dictionary: dict) -> bool:
    for value in dictionary.values():
        if isinstance(value, float) and math.isnan(value):
            return True
    return False


def is_nan(value) -> bool:
    if type(value) != float or type(value) != int:
        return value == 'nan'
    return math.isnan(value)


def group_dict_list_by_key(dict_list: list, key_to_grp: str, grouped_value_key='grouped_vals'):
    _grouped_data = []
    for key, value in groupby(sorted(dict_list, key=itemgetter(key_to_grp)), lambda x: x[key_to_grp]):
        _grouped_data.append({
            key_to_grp: key,
            grouped_value_key: list(value)
        })
    return _grouped_data


def read_workbook(path: str, format_key=False) -> list:
    df = pd.read_excel(path)
    df = df.replace(u'\xa0', ' ', regex=True)
    df.replace('NaN', np.nan, inplace=True)
    print(f'{path}: ')
    read_data = [x for x in tqdm(df.to_dict(orient="records")) if not pd.isna(x)]
    return read_data if not format_key else format_dictionaries(read_data)


def read_csv(path: str, format_key=False) -> list:
    read_data = []
    print(f'{path}: ')
    with open(path, 'r') as csv_file:
        reader = csv.DictReader(csv_file)
        for row in tqdm(reader):
            read_data.append(row)

    # read_data = json.dumps(read_data)
    return read_data if not format_key else format_dictionaries(read_data)


def read_excel_file(path: str, format_key=False) -> list:
    if path.isspace():
        return []
    print('Reading from', end=': ')
    if os.path.splitext(path)[1] == '.csv':
        print('csv file')
        return read_csv(path=path, format_key=format_key)
    elif os.path.splitext(path)[1] == '.xlsx':
        print('excel workbook file')
        return read_workbook(path=path, format_key=format_key)
    else:
        print('nothing')
        return []


def is_json(data):
    try:
        json.loads(data)
        return True
    except ValueError:
        return False


def write_to_json_file(data_list: list, file_path: str) -> None:
    if file_path.isspace():
        return
    with open(file_path, 'w') as json_file:
        json.dump(data_list, json_file)


def read_from_json_file(file_path: str):
    if file_path.isspace():
        return
    with open(file_path, 'r') as json_file:
        # Load the JSON data
        return json.load(json_file)


def save_dict_to_excel_workbook_with_row_formatting(file_path: str, headers: list, rows: list) -> None:
    if file_path.isspace():
        return
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(headers)
    print(f'Writing new excel workbook at {file_path}:')
    for row in tqdm(rows):
        sheet.append(row)
    workbook.save(file_path)


def append_excel_workbook(file_path: str, rows: list, worksheet_name='Sheet') -> bool:
    # if file_path.isspace():
    #    return
    workbook = load_workbook(filename=file_path)
    worksheet = workbook[worksheet_name]
    last_row = worksheet.max_row + 1  # start appending after the current last row
    if last_row + len(rows) >= 1048572:
        return True

    print(f"Appending {len(rows)} row{'s' if len(rows) > 1 else ''} to excel workbook at {file_path}:")
    for row in tqdm(rows):
        for col_index, value in enumerate(row, start=1):  # assign values to the corresponding cells in the worksheet
            worksheet.cell(row=last_row, column=col_index, value=value)
        last_row += 1  # make sure the next last row is appended to
    workbook.save(file_path)
    return False


def convert_datetime_obj_to_str(datetime_obj: datetime, str_format='%Y-%m-%d %H:%M:%S') -> str:
    return datetime_obj.strftime(str_format)


def random_seed_shuffle(seed: int, og_list: list) -> None:
    random.seed(seed)
    random.shuffle(og_list)


def save_running_logs(info: str, saved_path: str):
    print(info)
    with open(saved_path, 'a') as f:
        f.write(f'{info}\n')


def save_model(model, saved_location: str, optimiser, final=False, epoch=0, loss=0):
    _save_model_state = {
        'model_state_dict': model.state_dict(),
        'optimizer_state_dict': optimiser.state_dict(),
    }
    if not final:
        _save_model_state = {**_save_model_state, **{
            'epoch': epoch,
            'loss': loss
        }}
    torch.save(_save_model_state, saved_location)
    print(f'Pytorch models state is saved to {saved_location}')


def get_formatted_today_str(twelve_h=False):
    return datetime.now().strftime('%I:%M %p %d/%m/%Y' if twelve_h else '%H:%M %d/%m/%Y')


def list_of_dicts_to_single_dict(list_of_dicts: list, key_name: str, value_name: str) -> dict:
    return {data[key_name]: data[value_name] for data in list_of_dicts}


def remove_duplicate_in_list(og_list: list, sort_list=False) -> list:
    og_list = list(set(og_list))
    non_duplicate_list = []
    for item in og_list:
        if lower_case_and_clear_white_space(item, to_regex=False) not in non_duplicate_list or len(
                non_duplicate_list) < 1:
            non_duplicate_list.append(item)
    return sorted(non_duplicate_list) if sort_list else non_duplicate_list
